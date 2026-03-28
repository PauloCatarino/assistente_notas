"""Leitura inicial de obras a partir de ficheiros Excel."""

from __future__ import annotations

import re
from pathlib import Path

from models.schemas import ObraExcel
from utils.helpers import calcular_hash_ficheiro, obter_metadados_ficheiro

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - fallback simples
    load_workbook = None


class ImportadorObras:
    """Responsavel por localizar ficheiros Excel e ler metadados de obra."""

    EXTENSOES_VALIDAS = {".xlsx", ".xlsm", ".xls"}
    PADRAO_FICHEIRO_OBRA = re.compile(
        r"^(?P<nome_base>.+?)_"
        r"(?P<num_encomenda_phc>\d{4})_"
        r"(?P<versao_obra>\d{2})_"
        r"(?P<ano_obra>\d{2})_"
        r"(?P<cliente_codigo>.+)$",
        re.IGNORECASE,
    )

    def __init__(self, pasta_origem: Path) -> None:
        """Guarda a pasta onde os ficheiros Excel serao procurados."""
        self.pasta_origem = Path(pasta_origem)

    def listar_ficheiros_excel(self, extensoes: set[str] | None = None) -> list[Path]:
        """Lista ficheiros Excel suportados dentro da pasta configurada.

        O parametro `extensoes` permite restringir a pesquisa.
        Nesta fase sera usado sobretudo para procurar apenas `.xlsm`.
        """
        if not self.pasta_origem.exists():
            return []

        extensoes_validas = {extensao.lower() for extensao in (extensoes or self.EXTENSOES_VALIDAS)}

        return sorted(
            [
                caminho
                for caminho in self.pasta_origem.iterdir()
                if caminho.is_file()
                and caminho.suffix.lower() in extensoes_validas
                and not caminho.name.startswith("~$")
            ]
        )

    def ler_metadados_obra(self, caminho_ficheiro: Path) -> ObraExcel:
        """Le informacao simples do ficheiro para representar uma obra.

        A logica desta fase tenta interpretar nomes como:
        `Lista_Material_0560_01_26_JF_VIVA.xlsm`

        Quando o padrao bate certo, extraimos:
        - nome_base: `Lista_Material`
        - referencia_obra: `0560_01_26_JF_VIVA`
        - num_encomenda_phc: `0560`
        - versao_obra: `01`
        - ano_obra: `26`
        - cliente_codigo: `JF_VIVA`

        Se o nome nao seguir o padrao, o projeto continua a funcionar
        com valores vazios e usa o nome do ficheiro como fallback.
        """
        if load_workbook is None:
            raise RuntimeError("Instale `openpyxl` para ler ficheiros Excel.")

        caminho_ficheiro = Path(caminho_ficheiro).resolve()
        nome_ficheiro, tamanho_ficheiro, data_ficheiro = obter_metadados_ficheiro(caminho_ficheiro)
        hash_ficheiro = calcular_hash_ficheiro(caminho_ficheiro)
        dados_nome = self._interpretar_nome_ficheiro(caminho_ficheiro.stem)

        workbook = load_workbook(filename=caminho_ficheiro, read_only=True, data_only=True)
        try:
            folhas_disponiveis = list(workbook.sheetnames)
        finally:
            workbook.close()

        referencia_obra = dados_nome["referencia_obra"]
        nome_obra = referencia_obra or caminho_ficheiro.stem

        return ObraExcel(
            codigo_obra=referencia_obra or caminho_ficheiro.stem,
            nome_obra=nome_obra,
            caminho_ficheiro=str(caminho_ficheiro),
            folhas_disponiveis=folhas_disponiveis,
            nome_ficheiro=nome_ficheiro,
            nome_base=dados_nome["nome_base"],
            referencia_obra=referencia_obra,
            num_encomenda_phc=dados_nome["num_encomenda_phc"],
            versao_obra=dados_nome["versao_obra"],
            ano_obra=dados_nome["ano_obra"],
            cliente_codigo=dados_nome["cliente_codigo"],
            hash_ficheiro=hash_ficheiro,
            tamanho_ficheiro=tamanho_ficheiro,
            data_ficheiro=data_ficheiro,
        )

    def importar_pasta(self, extensoes: set[str] | None = None) -> list[ObraExcel]:
        """Importa os metadados basicos de todos os ficheiros Excel encontrados."""
        obras: list[ObraExcel] = []

        for caminho_ficheiro in self.listar_ficheiros_excel(extensoes=extensoes):
            try:
                obras.append(self.ler_metadados_obra(caminho_ficheiro))
            except Exception as erro:
                print(f"Erro ao ler obra '{caminho_ficheiro.name}': {erro}")

        return obras

    def _interpretar_nome_ficheiro(self, nome_sem_extensao: str) -> dict[str, str]:
        """Interpreta o nome do ficheiro e devolve partes estruturadas.

        Esta funcao e deliberadamente conservadora.
        Se o padrao nao for reconhecido, nao inventa valores.
        """
        correspondencia = self.PADRAO_FICHEIRO_OBRA.match(nome_sem_extensao)
        if not correspondencia:
            return {
                "nome_base": nome_sem_extensao,
                "referencia_obra": "",
                "num_encomenda_phc": "",
                "versao_obra": "",
                "ano_obra": "",
                "cliente_codigo": "",
            }

        dados = correspondencia.groupdict()
        referencia_obra = (
            f"{dados['num_encomenda_phc']}_{dados['versao_obra']}_{dados['ano_obra']}_{dados['cliente_codigo']}"
        )

        return {
            "nome_base": dados["nome_base"],
            "referencia_obra": referencia_obra,
            "num_encomenda_phc": dados["num_encomenda_phc"],
            "versao_obra": dados["versao_obra"],
            "ano_obra": dados["ano_obra"],
            "cliente_codigo": dados["cliente_codigo"],
        }
