"""Leitura de folhas Excel com mapeamento centralizado."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from config.excel_mapeamentos import ConfiguracaoFolhaExcel, FOLHAS_IMPORTACAO_EXCEL
from models.schemas import LinhaObra, ResultadoLeituraFolhaExcel
from utils.helpers import converter_para_numero, limpar_texto, normalizar_texto_chave

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - fallback simples
    load_workbook = None


class ImportadorFolhasExcel:
    """Lê as folhas de interesse e devolve linhas normalizadas para inserção."""

    CAMPOS_NUMERICOS = {"comp", "larg", "esp", "qt", "esp_mat", "esp_final"}

    def __init__(
        self,
        configuracoes_folhas: tuple[ConfiguracaoFolhaExcel, ...] = FOLHAS_IMPORTACAO_EXCEL,
    ) -> None:
        """Guarda a lista de folhas que devem ser importadas."""
        self.configuracoes_folhas = configuracoes_folhas

    def ler_folhas_configuradas(self, caminho_ficheiro: Path) -> list[ResultadoLeituraFolhaExcel]:
        """Lê todas as folhas configuradas do ficheiro Excel."""
        if load_workbook is None:
            raise RuntimeError("Instale `openpyxl` para ler ficheiros Excel.")

        workbook = load_workbook(filename=caminho_ficheiro, read_only=True, data_only=True)

        try:
            resultados: list[ResultadoLeituraFolhaExcel] = []

            for configuracao_folha in self.configuracoes_folhas:
                resultados.append(self._ler_folha(workbook, configuracao_folha))

            return resultados
        finally:
            workbook.close()

    def _ler_folha(self, workbook: Any, configuracao_folha: ConfiguracaoFolhaExcel) -> ResultadoLeituraFolhaExcel:
        """Lê uma folha Excel específica usando a configuração recebida."""
        nome_folha_real = self._resolver_nome_real_folha(workbook, configuracao_folha)
        worksheet = workbook[nome_folha_real]

        linha_cabecalho, cabecalhos_brutos = self._localizar_linha_cabecalho(
            worksheet,
            configuracao_folha,
        )

        mapa_campos = configuracao_folha.mapa_colunas_normalizado()
        indices_por_campo: dict[str, int] = {}
        colunas_mapeadas: dict[str, str] = {}

        for indice, cabecalho in enumerate(cabecalhos_brutos):
            chave_normalizada = normalizar_texto_chave(cabecalho)
            if not chave_normalizada:
                continue

            if chave_normalizada in mapa_campos and mapa_campos[chave_normalizada] not in indices_por_campo:
                campo = mapa_campos[chave_normalizada]
                indices_por_campo[campo] = indice
                colunas_mapeadas[campo] = limpar_texto(cabecalho)

        avisos: list[str] = []
        campos_em_falta = [
            campo
            for campo in mapa_campos.values()
            if campo not in indices_por_campo
        ]

        for campo in campos_em_falta:
            avisos.append(
                f"Aviso: a coluna esperada para o campo '{campo}' não foi encontrada na folha '{nome_folha_real}'."
            )

        linhas: list[LinhaObra] = []

        for numero_linha_excel, valores in enumerate(
            worksheet.iter_rows(min_row=linha_cabecalho + 1, values_only=True),
            start=linha_cabecalho + 1,
        ):
            if not any(valor is not None and str(valor).strip() for valor in valores):
                continue

            dados_brutos = self._extrair_dados_brutos(cabecalhos_brutos, valores)
            dados_mapeados = self._extrair_campos_mapeados(indices_por_campo, valores)

            linhas.append(
                LinhaObra(
                    obra_id=None,
                    estado_origem=configuracao_folha.estado_origem,
                    nome_folha_origem=nome_folha_real,
                    linha_excel=numero_linha_excel,
                    referencia=limpar_texto(dados_mapeados.get("ref_cliente") or dados_mapeados.get("artigo")),
                    designacao=limpar_texto(dados_mapeados.get("descricao")),
                    descricao=limpar_texto(dados_mapeados.get("descricao")),
                    material=limpar_texto(dados_mapeados.get("material")),
                    comp=converter_para_numero(dados_mapeados.get("comp")),
                    larg=converter_para_numero(dados_mapeados.get("larg")),
                    esp=converter_para_numero(dados_mapeados.get("esp")),
                    qt=converter_para_numero(dados_mapeados.get("qt")),
                    veio=limpar_texto(dados_mapeados.get("veio")),
                    orla_esq=limpar_texto(dados_mapeados.get("orla_esq")),
                    orla_dir=limpar_texto(dados_mapeados.get("orla_dir")),
                    orla_cima=limpar_texto(dados_mapeados.get("orla_cima")),
                    orla_baixo=limpar_texto(dados_mapeados.get("orla_baixo")),
                    cnc_1_raw=limpar_texto(dados_mapeados.get("cnc_1_raw")),
                    cnc_2_raw=limpar_texto(dados_mapeados.get("cnc_2_raw")),
                    enc=limpar_texto(dados_mapeados.get("enc")),
                    cliente=limpar_texto(dados_mapeados.get("cliente")),
                    ref_cliente=limpar_texto(dados_mapeados.get("ref_cliente")),
                    processo=limpar_texto(dados_mapeados.get("processo")),
                    artigo=limpar_texto(dados_mapeados.get("artigo")),
                    notas=limpar_texto(dados_mapeados.get("notas")),
                    id_linha_excel=limpar_texto(dados_mapeados.get("id_linha_excel")),
                    esp_mat=converter_para_numero(dados_mapeados.get("esp_mat")),
                    esp_final=converter_para_numero(dados_mapeados.get("esp_final")),
                    dados_brutos=dados_brutos,
                )
            )

        return ResultadoLeituraFolhaExcel(
            nome_folha_esperada=configuracao_folha.nome_logico,
            nome_folha_real=nome_folha_real,
            estado_origem=configuracao_folha.estado_origem,
            linha_cabecalho=linha_cabecalho,
            colunas_mapeadas=colunas_mapeadas,
            linhas=linhas,
            avisos=avisos,
        )

    def _resolver_nome_real_folha(self, workbook: Any, configuracao_folha: ConfiguracaoFolhaExcel) -> str:
        """Resolve o nome real da folha no ficheiro Excel."""
        nomes_aceites = {
            normalizar_texto_chave(nome)
            for nome in configuracao_folha.nomes_folha_aceites
        }

        for nome_folha in workbook.sheetnames:
            if normalizar_texto_chave(nome_folha) in nomes_aceites:
                return nome_folha

        raise ValueError(
            f"Folha não encontrada: esperava uma de {configuracao_folha.nomes_folha_aceites}."
        )

    def _localizar_linha_cabecalho(
        self,
        worksheet: Any,
        configuracao_folha: ConfiguracaoFolhaExcel,
    ) -> tuple[int, tuple[Any, ...]]:
        """Procura a linha de cabeçalho com base nas colunas esperadas."""
        mapa_campos = configuracao_folha.mapa_colunas_normalizado()
        cabecalhos_esperados = set(mapa_campos.keys())

        melhor_linha = 0
        maior_correspondencia = 0
        melhores_cabecalhos: tuple[Any, ...] = ()

        for numero_linha, valores in enumerate(
            worksheet.iter_rows(
                min_row=1,
                max_row=configuracao_folha.linha_procura_cabecalho_maxima,
                values_only=True,
            ),
            start=1,
        ):
            cabecalhos_normalizados = {
                normalizar_texto_chave(valor)
                for valor in valores
                if normalizar_texto_chave(valor)
            }

            correspondencias = len(cabecalhos_esperados.intersection(cabecalhos_normalizados))

            if correspondencias > maior_correspondencia:
                melhor_linha = numero_linha
                maior_correspondencia = correspondencias
                melhores_cabecalhos = tuple(valores)

        if melhor_linha == 0:
            raise ValueError(
                f"Não foi possível localizar o cabeçalho da folha '{worksheet.title}'."
            )

        return melhor_linha, melhores_cabecalhos

    @staticmethod
    def _extrair_dados_brutos(cabecalhos_brutos: tuple[Any, ...], valores: tuple[Any, ...]) -> dict[str, Any]:
        """Cria um dicionário simples com os dados brutos da linha."""
        dados_brutos: dict[str, Any] = {}

        for indice, valor in enumerate(valores):
            if indice >= len(cabecalhos_brutos):
                continue

            cabecalho = cabecalhos_brutos[indice]
            chave = normalizar_texto_chave(cabecalho)
            if not chave:
                continue

            dados_brutos[chave] = valor

        return dados_brutos

    def _extrair_campos_mapeados(
        self,
        indices_por_campo: dict[str, int],
        valores: tuple[Any, ...],
    ) -> dict[str, Any]:
        """Extrai apenas os campos relevantes para a base de dados."""
        dados_mapeados: dict[str, Any] = {}

        for campo, indice in indices_por_campo.items():
            if indice < len(valores):
                dados_mapeados[campo] = valores[indice]
            else:
                dados_mapeados[campo] = None

        return dados_mapeados


# Mantém o nome antigo disponível para compatibilidade simples.
ImportadorExcelCutRite = ImportadorFolhasExcel
