"""Servico simples para sugestoes, validacao manual e feedback."""

from __future__ import annotations

import csv
import hashlib
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from config.settings import ConfiguracaoProjeto, carregar_configuracoes
from config.sugestoes_notas import (
    AJUSTE_MAXIMO_NEGATIVO,
    AJUSTE_MAXIMO_POSITIVO,
    COLUNAS_EXPORTACAO_VALIDACAO,
    LIMIAR_CANDIDATO,
    LIMIAR_SUGESTAO_BASE,
    LIMIAR_SUGESTAO_RECALIBRADA,
    MIN_RESPOSTAS_RECALIBRAR_DESCRICAO_NOTA,
    MIN_RESPOSTAS_RECALIBRAR_NOTA,
    VALIDACOES_ACEITES,
    VALIDACOES_EDITADAS,
    VALIDACOES_REJEITADAS,
)
from database.db_connection import GestorLigacaoMySQL
from database.repositorio_importacao_excel import RepositorioImportacaoExcel
from database.repositorio_sugestoes import RepositorioSugestoes
from importers.importar_excel_cutrite import ImportadorFolhasExcel
from importers.importar_obras import ImportadorObras
from models.schemas import (
    FeedbackSugestaoNota,
    LinhaObra,
    RelatorioQualidadeSugestoes,
    ResultadoAnaliseSugestoesExcel,
    ResultadoImportacaoFeedbackSugestoes,
    SugestaoNotaLinhaExcel,
)
from utils.chave_ligacao import normalizar_numero_para_comparacao, normalizar_texto_para_comparacao
from utils.helpers import converter_para_numero, limpar_texto, normalizar_texto_chave

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font
except ImportError:  # pragma: no cover - fallback simples
    Workbook = None
    Font = None
    load_workbook = None


class ServicoSugestoes:
    """Centraliza o ciclo pratico de validacao e recalibracao."""

    PESOS_CAMPOS = {
        "descricao": 25,
        "material": 20,
        "artigo": 15,
        "veio": 5,
        "esp": 10,
        "esp_mat": 5,
        "esp_final": 5,
        "orla_esq": 5,
        "orla_dir": 5,
        "orla_cima": 3,
        "orla_baixo": 2,
    }

    def __init__(self, configuracao: ConfiguracaoProjeto | None = None) -> None:
        """Prepara dependencias do motor de sugestoes."""
        self.configuracao = configuracao or carregar_configuracoes()
        self.importador_obras = ImportadorObras(self.configuracao.base_dir)
        self.importador_folhas = ImportadorFolhasExcel()
        self.gestor_bd = GestorLigacaoMySQL(self.configuracao)
        self.repositorio_importacao = RepositorioImportacaoExcel()
        self.repositorio_sugestoes = RepositorioSugestoes()

    def analisar_ficheiro_excel(
        self,
        caminho_ficheiro: Path,
        caminho_csv_saida: Path | None = None,
        caminho_excel_saida: Path | None = None,
        gerar_csv: bool = False,
        gerar_excel: bool = True,
        usar_recalibracao: bool = True,
    ) -> ResultadoAnaliseSugestoesExcel:
        """Analisa um ficheiro Excel e gera ficheiros de validacao."""
        caminho_ficheiro = Path(caminho_ficheiro).resolve()
        if not caminho_ficheiro.exists():
            raise FileNotFoundError(f"Ficheiro nao encontrado: {caminho_ficheiro}")

        obra_excel = self.importador_obras.ler_metadados_obra(caminho_ficheiro)
        resultados_folhas = self.importador_folhas.ler_folhas_configuradas(caminho_ficheiro)
        resultado_folha = self._selecionar_folha_para_sugestoes(resultados_folhas)

        conexao = self.gestor_bd.criar_conexao()
        if conexao is None:
            raise RuntimeError(
                "Nao foi possivel criar ligacao ao MySQL. Verifique o ficheiro .env e a disponibilidade do servidor."
            )

        feedbacks_reais: list[FeedbackSugestaoNota] = []

        try:
            self.repositorio_importacao.preparar_estrutura_importacao(conexao)
            self.repositorio_sugestoes.preparar_estrutura_feedback(conexao)

            obra_existente = self.repositorio_importacao.procurar_obra_existente(conexao, obra_excel)
            obra_id_excluir = int(obra_existente["id"]) if obra_existente else None

            total_obras_historico = self.repositorio_sugestoes.contar_obras(conexao)
            total_linhas_com_nota_historica = self.repositorio_sugestoes.contar_linhas_com_nota_historica(
                conexao,
                excluir_obra_id=obra_id_excluir,
            )
            historico = self.repositorio_sugestoes.listar_linhas_historicas_com_nota(
                conexao,
                excluir_obra_id=obra_id_excluir,
            )

            if usar_recalibracao:
                feedbacks_reais = self.repositorio_sugestoes.listar_feedback(conexao)
        finally:
            self.gestor_bd.fechar_conexao(conexao)

        if obra_id_excluir is not None and total_obras_historico > 0:
            total_obras_historico -= 1

        contexto_recalibracao = self.construir_contexto_recalibracao(feedbacks_reais) if usar_recalibracao else {}

        sugestoes = self.gerar_sugestoes_para_linhas(
            obra_id=obra_id_excluir,
            linhas_alvo=resultado_folha.linhas,
            historico=historico,
            usar_recalibracao=usar_recalibracao,
            contexto_recalibracao=contexto_recalibracao,
        )

        if caminho_csv_saida is None and gerar_csv:
            caminho_csv_saida = self.configuracao.logs_dir / f"validacao_sugestoes_{caminho_ficheiro.stem}.csv"
        if caminho_excel_saida is None and gerar_excel:
            caminho_excel_saida = self.configuracao.logs_dir / f"validacao_sugestoes_{caminho_ficheiro.stem}.xlsx"

        caminho_csv_final = ""
        caminho_excel_final = ""

        if gerar_csv and caminho_csv_saida is not None:
            caminho_csv_saida = Path(caminho_csv_saida).resolve()
            self.exportar_csv_validacao(caminho_csv_saida, sugestoes)
            caminho_csv_final = str(caminho_csv_saida)

        if gerar_excel and caminho_excel_saida is not None:
            caminho_excel_saida = Path(caminho_excel_saida).resolve()
            self.exportar_excel_validacao(caminho_excel_saida, sugestoes)
            caminho_excel_final = str(caminho_excel_saida)

        total_sugestoes_geradas = sum(1 for sugestao in sugestoes if sugestao.sugestao_1)
        total_linhas_sem_sugestao = len(sugestoes) - total_sugestoes_geradas

        return ResultadoAnaliseSugestoesExcel(
            caminho_ficheiro=str(caminho_ficheiro),
            nome_folha_analisada=resultado_folha.nome_folha_real,
            obra_id=obra_id_excluir,
            total_obras_historico=total_obras_historico,
            total_linhas_com_nota_historica=total_linhas_com_nota_historica,
            total_linhas_analisadas=len(sugestoes),
            total_sugestoes_geradas=total_sugestoes_geradas,
            total_linhas_sem_sugestao=total_linhas_sem_sugestao,
            caminho_csv_saida=caminho_csv_final,
            caminho_excel_saida=caminho_excel_final,
            sugestoes=sugestoes,
        )

    def importar_feedback_validacao(self, caminho_ficheiro: Path) -> ResultadoImportacaoFeedbackSugestoes:
        """Importa feedback preenchido num CSV ou Excel de validacao."""
        caminho_ficheiro = Path(caminho_ficheiro).resolve()
        if not caminho_ficheiro.exists():
            raise FileNotFoundError(f"Ficheiro nao encontrado: {caminho_ficheiro}")

        feedbacks_lidos = self._ler_feedback_de_ficheiro(caminho_ficheiro)
        feedbacks_validos = [
            feedback
            for feedback in feedbacks_lidos
            if feedback.obra_id and feedback.linha_excel and self._feedback_tem_resposta_manual(feedback)
        ]
        total_ignorados = len(feedbacks_lidos) - len(feedbacks_validos)

        conexao = self.gestor_bd.criar_conexao()
        if conexao is None:
            raise RuntimeError(
                "Nao foi possivel criar ligacao ao MySQL. Verifique o ficheiro .env e a disponibilidade do servidor."
            )

        total_duplicados = 0
        feedbacks_para_gravar: list[FeedbackSugestaoNota] = []

        try:
            self.repositorio_importacao.preparar_estrutura_importacao(conexao)
            self.repositorio_sugestoes.preparar_estrutura_feedback(conexao)

            obra_ids = {feedback.obra_id for feedback in feedbacks_validos}
            feedback_existente = self.repositorio_sugestoes.listar_feedback_indexado(conexao, obra_ids=obra_ids)

            for feedback in feedbacks_validos:
                feedback.feedback_hash = self.criar_feedback_hash(feedback)
                chave_feedback = (feedback.obra_id, feedback.linha_excel)
                existente = feedback_existente.get(chave_feedback)

                if existente is not None:
                    hash_existente = existente.feedback_hash or self.criar_feedback_hash(existente)
                    if hash_existente == feedback.feedback_hash:
                        total_duplicados += 1
                        continue

                feedbacks_para_gravar.append(feedback)
                feedback_existente[chave_feedback] = feedback

            total_importado = self.repositorio_sugestoes.guardar_feedback(conexao, feedbacks_para_gravar)
            conexao.commit()
        except Exception:
            conexao.rollback()
            raise
        finally:
            self.gestor_bd.fechar_conexao(conexao)

        return ResultadoImportacaoFeedbackSugestoes(
            caminho_ficheiro=str(caminho_ficheiro),
            total_registos_lidos=len(feedbacks_lidos),
            total_registos_importados=total_importado,
            total_registos_ignorados=total_ignorados,
            total_registos_duplicados=total_duplicados,
        )

    def gerar_relatorio_qualidade(self, obra_id: int | None = None) -> RelatorioQualidadeSugestoes:
        """Gera um relatorio simples com base no feedback guardado."""
        conexao = self.gestor_bd.criar_conexao()
        if conexao is None:
            raise RuntimeError(
                "Nao foi possivel criar ligacao ao MySQL. Verifique o ficheiro .env e a disponibilidade do servidor."
            )

        try:
            self.repositorio_sugestoes.preparar_estrutura_feedback(conexao)
            feedbacks = self.repositorio_sugestoes.listar_feedback(conexao, obra_id=obra_id)
        finally:
            self.gestor_bd.fechar_conexao(conexao)

        contexto_recalibracao = self.construir_contexto_recalibracao(feedbacks)
        return self.construir_relatorio_qualidade(feedbacks, contexto_recalibracao=contexto_recalibracao)

    def gerar_sugestoes_para_linhas(
        self,
        obra_id: int | None,
        linhas_alvo: list[LinhaObra],
        historico: list[LinhaObra],
        usar_recalibracao: bool = True,
        contexto_recalibracao: dict[str, Any] | None = None,
    ) -> list[SugestaoNotaLinhaExcel]:
        """Gera sugestoes simples para cada linha lida do Excel."""
        sugestoes: list[SugestaoNotaLinhaExcel] = []
        contexto = contexto_recalibracao or {}

        for linha_alvo in linhas_alvo:
            sugestoes.append(
                self._gerar_sugestao_para_linha(
                    obra_id,
                    linha_alvo,
                    historico,
                    usar_recalibracao=usar_recalibracao,
                    contexto_recalibracao=contexto,
                )
            )

        return sugestoes

    def construir_relatorio_qualidade(
        self,
        feedbacks: list[FeedbackSugestaoNota],
        contexto_recalibracao: dict[str, Any] | None = None,
    ) -> RelatorioQualidadeSugestoes:
        """Constroi o relatorio de qualidade a partir do feedback carregado."""
        feedbacks_considerados = [feedback for feedback in feedbacks if self._feedback_tem_resposta_manual(feedback)]
        total_linhas_analisadas = len(feedbacks_considerados)
        total_linhas_com_sugestao = sum(1 for feedback in feedbacks_considerados if feedback.sugestao_1.strip())
        total_linhas_sem_sugestao = total_linhas_analisadas - total_linhas_com_sugestao

        estados_feedback = [
            (
                feedback,
                self.normalizar_estado_feedback(
                    feedback.validacao_utilizador,
                    feedback.sugestao_1,
                    feedback.nota_final_utilizador,
                ),
            )
            for feedback in feedbacks_considerados
        ]

        total_sugestoes_aceites = sum(1 for _, estado in estados_feedback if estado == "ACEITE")
        total_sugestoes_rejeitadas = sum(1 for _, estado in estados_feedback if estado == "REJEITADA")
        total_sugestoes_editadas = sum(1 for _, estado in estados_feedback if estado == "EDITADA")

        taxa_cobertura = self._calcular_percentagem(total_linhas_com_sugestao, total_linhas_analisadas)
        taxa_aceitacao = self._calcular_percentagem(total_sugestoes_aceites, total_linhas_com_sugestao)
        taxa_rejeicao = self._calcular_percentagem(total_sugestoes_rejeitadas, total_linhas_com_sugestao)

        descricoes_acertos = Counter(
            feedback.descricao
            for feedback, estado in estados_feedback
            if estado == "ACEITE" and feedback.descricao
        )
        descricoes_falhas = Counter(
            feedback.descricao
            for feedback, estado in estados_feedback
            if estado in {"REJEITADA", "EDITADA"} and feedback.descricao
        )
        notas_aceites = Counter(
            feedback.sugestao_1
            for feedback, estado in estados_feedback
            if estado == "ACEITE" and feedback.sugestao_1
        )
        notas_rejeitadas = Counter(
            feedback.sugestao_1
            for feedback, estado in estados_feedback
            if estado == "REJEITADA" and feedback.sugestao_1
        )

        contexto = contexto_recalibracao or self.construir_contexto_recalibracao(feedbacks)
        totais_recalibrados = self._estimar_metricas_recalibradas(estados_feedback, contexto)

        return RelatorioQualidadeSugestoes(
            total_linhas_analisadas=total_linhas_analisadas,
            total_linhas_com_sugestao=total_linhas_com_sugestao,
            total_linhas_sem_sugestao=total_linhas_sem_sugestao,
            total_sugestoes_aceites=total_sugestoes_aceites,
            total_sugestoes_rejeitadas=total_sugestoes_rejeitadas,
            total_sugestoes_editadas=total_sugestoes_editadas,
            taxa_cobertura=taxa_cobertura,
            taxa_aceitacao=taxa_aceitacao,
            taxa_rejeicao=taxa_rejeicao,
            cobertura_recalibrada_estimada=totais_recalibrados["cobertura_recalibrada_estimada"],
            aceitacao_recalibrada_estimada=totais_recalibrados["aceitacao_recalibrada_estimada"],
            rejeicao_recalibrada_estimada=totais_recalibrados["rejeicao_recalibrada_estimada"],
            total_padroes_recalibrados=totais_recalibrados["total_padroes_recalibrados"],
            top_descricoes_com_mais_acertos=descricoes_acertos.most_common(5),
            top_descricoes_com_mais_falhas=descricoes_falhas.most_common(5),
            top_notas_mais_aceites=notas_aceites.most_common(5),
            top_notas_mais_rejeitadas=notas_rejeitadas.most_common(5),
            top_sugestoes_penalizadas=totais_recalibrados["top_sugestoes_penalizadas"],
            top_sugestoes_reforcadas=totais_recalibrados["top_sugestoes_reforcadas"],
        )

    def construir_contexto_recalibracao(
        self,
        feedbacks: list[FeedbackSugestaoNota],
    ) -> dict[str, Any]:
        """Resume o feedback real em ajustes simples e transparentes."""
        estatisticas_por_nota: dict[str, dict[str, int]] = defaultdict(self._criar_estatisticas_feedback)
        estatisticas_por_descricao_nota: dict[tuple[str, str], dict[str, int]] = defaultdict(
            self._criar_estatisticas_feedback
        )
        rotulos_notas: dict[str, str] = {}
        total_feedback_valido = 0

        for feedback in feedbacks:
            nota_normalizada = normalizar_texto_para_comparacao(feedback.sugestao_1)
            if not nota_normalizada:
                continue

            estado = self.normalizar_estado_feedback(
                feedback.validacao_utilizador,
                feedback.sugestao_1,
                feedback.nota_final_utilizador,
            )
            if estado == "SEM_RESPOSTA":
                continue

            descricao_normalizada = normalizar_texto_para_comparacao(feedback.descricao)
            rotulos_notas.setdefault(nota_normalizada, feedback.sugestao_1)

            self._acumular_estado_estatistico(estatisticas_por_nota[nota_normalizada], estado)
            if descricao_normalizada:
                self._acumular_estado_estatistico(
                    estatisticas_por_descricao_nota[(descricao_normalizada, nota_normalizada)],
                    estado,
                )
            total_feedback_valido += 1

        ajustes_por_nota: dict[str, int] = {}
        ajustes_por_descricao_nota: dict[tuple[str, str], int] = {}

        for nota_normalizada, estatisticas in estatisticas_por_nota.items():
            ajuste = self._calcular_ajuste_estatistico(
                estatisticas,
                minimo_respostas=MIN_RESPOSTAS_RECALIBRAR_NOTA,
                peso_aceite=4,
                peso_rejeitada=7,
                peso_editada=4,
                bonus_aceite=4,
                bonus_falha=4,
                limite_inferior=AJUSTE_MAXIMO_NEGATIVO,
                limite_superior=AJUSTE_MAXIMO_POSITIVO,
            )
            if ajuste != 0:
                ajustes_por_nota[nota_normalizada] = ajuste

        for chave_descricao_nota, estatisticas in estatisticas_por_descricao_nota.items():
            ajuste = self._calcular_ajuste_estatistico(
                estatisticas,
                minimo_respostas=MIN_RESPOSTAS_RECALIBRAR_DESCRICAO_NOTA,
                peso_aceite=3,
                peso_rejeitada=5,
                peso_editada=2,
                bonus_aceite=2,
                bonus_falha=2,
                limite_inferior=-12,
                limite_superior=8,
            )
            if ajuste != 0:
                ajustes_por_descricao_nota[chave_descricao_nota] = ajuste

        return {
            "ativo": bool(ajustes_por_nota or ajustes_por_descricao_nota),
            "total_feedback_valido": total_feedback_valido,
            "estatisticas_por_nota": dict(estatisticas_por_nota),
            "ajustes_por_nota": ajustes_por_nota,
            "ajustes_por_descricao_nota": ajustes_por_descricao_nota,
            "rotulos_notas": rotulos_notas,
        }

    def calcular_score_sugestao(self, linha_alvo: LinhaObra, linha_historica: LinhaObra) -> tuple[int, list[str]]:
        """Calcula um score simples entre uma linha alvo e uma linha historica."""
        score = 0
        campos_coincidentes: list[str] = []

        for campo_texto in ("descricao", "material", "artigo", "veio", "orla_esq", "orla_dir", "orla_cima", "orla_baixo"):
            score, campos_coincidentes = self._somar_score_texto(
                linha_alvo,
                linha_historica,
                campo_texto,
                score,
                campos_coincidentes,
            )

        for campo_numero in ("esp", "esp_mat", "esp_final"):
            score, campos_coincidentes = self._somar_score_numero(
                linha_alvo,
                linha_historica,
                campo_numero,
                score,
                campos_coincidentes,
            )

        campos_principais = {"descricao", "material", "artigo"}
        if not campos_principais.intersection(campos_coincidentes):
            return 0, []

        return min(score, 100), campos_coincidentes

    def criar_feedback_hash(self, feedback: FeedbackSugestaoNota) -> str:
        """Cria um hash estavel para detetar o mesmo feedback importado novamente."""
        componentes = [
            f"obra_id={feedback.obra_id}",
            f"linha_excel={feedback.linha_excel}",
            f"sugestao_1={normalizar_texto_para_comparacao(feedback.sugestao_1)}",
            f"score_1={int(feedback.score_1 or 0)}",
            f"sugestao_2={normalizar_texto_para_comparacao(feedback.sugestao_2)}",
            f"score_2={int(feedback.score_2 or 0)}",
            f"validacao={normalizar_texto_para_comparacao(feedback.validacao_utilizador)}",
            f"nota_final={normalizar_texto_para_comparacao(feedback.nota_final_utilizador)}",
        ]
        return hashlib.sha256("|".join(componentes).encode("utf-8")).hexdigest()

    def exportar_csv_validacao(
        self,
        caminho_csv_saida: Path,
        sugestoes: list[SugestaoNotaLinhaExcel],
    ) -> None:
        """Exporta um CSV com colunas prontas para validacao manual."""
        caminho_csv_saida = Path(caminho_csv_saida)
        caminho_csv_saida.parent.mkdir(parents=True, exist_ok=True)

        with caminho_csv_saida.open("w", encoding="utf-8-sig", newline="") as ficheiro_csv:
            escritor = csv.writer(ficheiro_csv, delimiter=";")
            escritor.writerow(COLUNAS_EXPORTACAO_VALIDACAO)

            for sugestao in sugestoes:
                escritor.writerow(self._linha_para_exportacao(sugestao))

    def exportar_excel_validacao(
        self,
        caminho_excel_saida: Path,
        sugestoes: list[SugestaoNotaLinhaExcel],
    ) -> None:
        """Exporta um ficheiro Excel simples para validacao manual."""
        if Workbook is None or Font is None:
            raise RuntimeError("Instale `openpyxl` para gerar o ficheiro Excel de validacao.")

        caminho_excel_saida = Path(caminho_excel_saida)
        caminho_excel_saida.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        folha = workbook.active
        folha.title = "validacao_sugestoes"
        folha.append(list(COLUNAS_EXPORTACAO_VALIDACAO))

        for celula in folha[1]:
            celula.font = Font(bold=True)

        for sugestao in sugestoes:
            folha.append(self._linha_para_exportacao(sugestao))

        folha.freeze_panes = "A2"
        folha.auto_filter.ref = folha.dimensions

        larguras = {
            "A": 10,
            "B": 12,
            "C": 30,
            "D": 40,
            "E": 20,
            "F": 20,
            "G": 40,
            "H": 10,
            "I": 40,
            "J": 10,
            "K": 60,
            "L": 18,
            "M": 30,
        }
        for coluna, largura in larguras.items():
            folha.column_dimensions[coluna].width = largura

        workbook.save(caminho_excel_saida)
        workbook.close()

    def normalizar_estado_feedback(
        self,
        validacao_utilizador: str,
        sugestao_1: str,
        nota_final_utilizador: str,
    ) -> str:
        """Normaliza o feedback manual para estados simples e comparaveis."""
        validacao_normalizada = normalizar_texto_para_comparacao(validacao_utilizador)
        sugestao_normalizada = normalizar_texto_para_comparacao(sugestao_1)
        nota_final_normalizada = normalizar_texto_para_comparacao(nota_final_utilizador)

        if validacao_normalizada in VALIDACOES_ACEITES:
            return "ACEITE"
        if validacao_normalizada in VALIDACOES_REJEITADAS:
            return "REJEITADA"
        if validacao_normalizada in VALIDACOES_EDITADAS:
            return "EDITADA"

        if "nao aceite" in validacao_normalizada:
            return "REJEITADA"
        if "aceite sugestao 1" in validacao_normalizada:
            return "ACEITE"
        if "aceite sugestao 2" in validacao_normalizada:
            return "EDITADA"

        if sugestao_normalizada and nota_final_normalizada:
            if sugestao_normalizada == nota_final_normalizada:
                return "ACEITE"
            return "EDITADA"

        return "SEM_RESPOSTA"

    def _gerar_sugestao_para_linha(
        self,
        obra_id: int | None,
        linha_alvo: LinhaObra,
        historico: list[LinhaObra],
        usar_recalibracao: bool,
        contexto_recalibracao: dict[str, Any],
    ) -> SugestaoNotaLinhaExcel:
        """Gera as duas melhores sugestoes para uma linha."""
        candidatos_por_nota: dict[str, dict[str, Any]] = defaultdict(
            lambda: {
                "melhor_score": 0,
                "ocorrencias": 0,
                "campos": [],
            }
        )

        for linha_historica in historico:
            score, campos_coincidentes = self.calcular_score_sugestao(linha_alvo, linha_historica)
            if score < LIMIAR_CANDIDATO:
                continue

            nota_historica = (linha_historica.notas or "").strip()
            if not nota_historica:
                continue

            agregado = candidatos_por_nota[nota_historica]
            agregado["ocorrencias"] += 1

            if score > agregado["melhor_score"]:
                agregado["melhor_score"] = score
                agregado["campos"] = campos_coincidentes

        sugestoes_ordenadas: list[tuple[str, int, int, list[str], int, int, str]] = []

        for nota_historica, dados in candidatos_por_nota.items():
            score_base = min(100, dados["melhor_score"] + min(10, (dados["ocorrencias"] - 1) * 5))
            ajuste_recalibracao = 0
            detalhe_recalibracao = ""

            if usar_recalibracao and contexto_recalibracao.get("ativo"):
                ajuste_recalibracao, detalhe_recalibracao = self._calcular_ajuste_recalibracao(
                    descricao=linha_alvo.descricao,
                    nota=nota_historica,
                    contexto_recalibracao=contexto_recalibracao,
                )

            score_final = max(0, min(100, score_base + ajuste_recalibracao))
            sugestoes_ordenadas.append(
                (
                    nota_historica,
                    score_final,
                    int(dados["ocorrencias"]),
                    list(dados["campos"]),
                    score_base,
                    ajuste_recalibracao,
                    detalhe_recalibracao,
                )
            )

        sugestoes_ordenadas.sort(key=lambda item: (-item[1], -item[2], item[0]))
        limiar_sugestao = self._obter_limiar_sugestao(usar_recalibracao, contexto_recalibracao)
        sugestoes_finais = [item for item in sugestoes_ordenadas if item[1] >= limiar_sugestao][:2]

        if not sugestoes_finais:
            justificacao = "Sem historico suficiente para sugerir nota com seguranca."
            if sugestoes_ordenadas:
                justificacao = "Existe historico, mas abaixo do limiar minimo de confianca."
                if usar_recalibracao and contexto_recalibracao.get("ativo"):
                    justificacao = (
                        "Existe historico, mas a recalibracao penalizou os candidatos "
                        "ou o score ficou abaixo do limiar mais exigente."
                    )

            return SugestaoNotaLinhaExcel(
                obra_id=obra_id,
                linha_excel=linha_alvo.linha_excel,
                descricao=linha_alvo.descricao,
                material=linha_alvo.material,
                artigo=linha_alvo.artigo,
                notas_atual=linha_alvo.notas,
                justificacao=justificacao,
            )

        nota_1, score_1, ocorrencias_1, campos_1, score_base_1, ajuste_1, detalhe_1 = sugestoes_finais[0]
        nota_2 = ""
        score_2 = 0

        if len(sugestoes_finais) > 1:
            nota_2, score_2, _, _, _, _, _ = sugestoes_finais[1]

        lista_campos = ", ".join(campos_1[:6]) if campos_1 else "sem campos fortes"
        justificacao = (
            f"{ocorrencias_1} linha(s) historica(s) semelhante(s); "
            f"campos coincidentes: {lista_campos}; "
            f"score base: {score_base_1}."
        )
        if ajuste_1 != 0:
            justificacao += f" ajuste por feedback: {ajuste_1:+d} ({detalhe_1})."

        return SugestaoNotaLinhaExcel(
            obra_id=obra_id,
            linha_excel=linha_alvo.linha_excel,
            descricao=linha_alvo.descricao,
            material=linha_alvo.material,
            artigo=linha_alvo.artigo,
            notas_atual=linha_alvo.notas,
            sugestao_1=nota_1,
            score_1=score_1,
            sugestao_2=nota_2,
            score_2=score_2,
            justificacao=justificacao,
        )

    def _ler_feedback_de_ficheiro(self, caminho_ficheiro: Path) -> list[FeedbackSugestaoNota]:
        """Le feedback a partir de CSV ou Excel gerado para validacao."""
        extensao = caminho_ficheiro.suffix.lower()

        if extensao == ".csv":
            return self._ler_feedback_csv(caminho_ficheiro)

        if extensao in {".xlsx", ".xlsm"}:
            return self._ler_feedback_excel(caminho_ficheiro)

        raise ValueError("Formato de feedback nao suportado. Use .csv, .xlsx ou .xlsm.")

    def _ler_feedback_csv(self, caminho_ficheiro: Path) -> list[FeedbackSugestaoNota]:
        """Le feedback manual a partir de CSV."""
        feedbacks: list[FeedbackSugestaoNota] = []

        with Path(caminho_ficheiro).open("r", encoding="utf-8-sig", newline="") as ficheiro_csv:
            leitor = csv.DictReader(ficheiro_csv, delimiter=";")
            for linha in leitor:
                feedback = self._criar_feedback_a_partir_de_linha(linha)
                if feedback is not None:
                    feedbacks.append(feedback)

        return feedbacks

    def _ler_feedback_excel(self, caminho_ficheiro: Path) -> list[FeedbackSugestaoNota]:
        """Le feedback manual a partir de um Excel de validacao."""
        if load_workbook is None:
            raise RuntimeError("Instale `openpyxl` para ler o ficheiro Excel de feedback.")

        workbook = load_workbook(filename=caminho_ficheiro, read_only=True, data_only=True)
        feedbacks: list[FeedbackSugestaoNota] = []

        try:
            folha = workbook.active
            linhas = folha.iter_rows(values_only=True)
            cabecalhos = next(linhas, None)
            if cabecalhos is None:
                return []

            chaves = [normalizar_texto_chave(cabecalho) for cabecalho in cabecalhos]

            for valores in linhas:
                dados_linha = {
                    chaves[indice]: valores[indice]
                    for indice in range(min(len(chaves), len(valores)))
                    if chaves[indice]
                }
                feedback = self._criar_feedback_a_partir_de_linha(dados_linha)
                if feedback is not None:
                    feedbacks.append(feedback)
        finally:
            workbook.close()

        return feedbacks

    def _criar_feedback_a_partir_de_linha(self, dados_linha: dict[str, Any]) -> FeedbackSugestaoNota | None:
        """Converte uma linha lida do ficheiro de validacao em feedback estruturado."""
        obra_id = int(converter_para_numero(dados_linha.get("obra_id")) or 0)
        linha_excel = int(converter_para_numero(dados_linha.get("linha_excel")) or 0)

        if obra_id <= 0 or linha_excel <= 0:
            return None

        return FeedbackSugestaoNota(
            obra_id=obra_id,
            linha_excel=linha_excel,
            descricao=limpar_texto(dados_linha.get("descricao")),
            material=limpar_texto(dados_linha.get("material")),
            artigo=limpar_texto(dados_linha.get("artigo")),
            notas_atual=limpar_texto(dados_linha.get("notas_atual")),
            sugestao_1=limpar_texto(dados_linha.get("sugestao_1")),
            score_1=int(converter_para_numero(dados_linha.get("score_1")) or 0),
            sugestao_2=limpar_texto(dados_linha.get("sugestao_2")),
            score_2=int(converter_para_numero(dados_linha.get("score_2")) or 0),
            justificacao=limpar_texto(dados_linha.get("justificacao")),
            validacao_utilizador=limpar_texto(dados_linha.get("validacao_utilizador")),
            nota_final_utilizador=limpar_texto(dados_linha.get("nota_final_utilizador")),
            data_feedback=datetime.now(),
        )

    def _feedback_tem_resposta_manual(self, feedback: FeedbackSugestaoNota) -> bool:
        """Indica se a linha tem resposta manual suficiente para entrar no ciclo."""
        if normalizar_texto_para_comparacao(feedback.validacao_utilizador):
            return True

        if normalizar_texto_para_comparacao(feedback.nota_final_utilizador):
            return True

        return False

    def _linha_para_exportacao(self, sugestao: SugestaoNotaLinhaExcel) -> list[Any]:
        """Converte a sugestao numa linha pronta para CSV ou Excel."""
        return [
            sugestao.obra_id or "",
            sugestao.linha_excel,
            sugestao.descricao,
            sugestao.material,
            sugestao.artigo,
            sugestao.notas_atual,
            sugestao.sugestao_1,
            sugestao.score_1,
            sugestao.sugestao_2,
            sugestao.score_2,
            sugestao.justificacao,
            sugestao.validacao_utilizador,
            sugestao.nota_final_utilizador,
        ]

    def _selecionar_folha_para_sugestoes(self, resultados_folhas: list[Any]) -> Any:
        """Escolhe a melhor folha para gerar sugestoes."""
        for resultado in resultados_folhas:
            if resultado.nome_folha_esperada == "LISTAGEM_CUT_RITE":
                return resultado

        for resultado in resultados_folhas:
            if resultado.linhas:
                return resultado

        raise ValueError("Nao foi encontrada nenhuma folha com linhas para analisar.")

    def _somar_score_texto(
        self,
        linha_alvo: LinhaObra,
        linha_historica: LinhaObra,
        campo: str,
        score_atual: int,
        campos_coincidentes: list[str],
    ) -> tuple[int, list[str]]:
        """Soma score quando um campo textual coincide apos normalizacao."""
        valor_alvo = normalizar_texto_para_comparacao(getattr(linha_alvo, campo, ""))
        valor_historico = normalizar_texto_para_comparacao(getattr(linha_historica, campo, ""))

        if valor_alvo and valor_alvo == valor_historico:
            return score_atual + self.PESOS_CAMPOS[campo], campos_coincidentes + [campo]

        return score_atual, campos_coincidentes

    def _somar_score_numero(
        self,
        linha_alvo: LinhaObra,
        linha_historica: LinhaObra,
        campo: str,
        score_atual: int,
        campos_coincidentes: list[str],
    ) -> tuple[int, list[str]]:
        """Soma score quando um campo numerico coincide dentro de tolerancia simples."""
        valor_alvo = normalizar_numero_para_comparacao(getattr(linha_alvo, campo, None))
        valor_historico = normalizar_numero_para_comparacao(getattr(linha_historica, campo, None))

        if valor_alvo is None or valor_historico is None:
            return score_atual, campos_coincidentes

        if abs(valor_alvo - valor_historico) <= 0.2:
            return score_atual + self.PESOS_CAMPOS[campo], campos_coincidentes + [campo]

        return score_atual, campos_coincidentes

    def _obter_limiar_sugestao(self, usar_recalibracao: bool, contexto_recalibracao: dict[str, Any]) -> int:
        """Escolhe o limiar correto para esta analise."""
        if usar_recalibracao and contexto_recalibracao.get("ativo"):
            return LIMIAR_SUGESTAO_RECALIBRADA
        return LIMIAR_SUGESTAO_BASE

    def _calcular_ajuste_recalibracao(
        self,
        descricao: str,
        nota: str,
        contexto_recalibracao: dict[str, Any],
    ) -> tuple[int, str]:
        """Calcula um ajuste final simples para a nota candidata."""
        nota_normalizada = normalizar_texto_para_comparacao(nota)
        descricao_normalizada = normalizar_texto_para_comparacao(descricao)

        ajuste_nota = int(contexto_recalibracao.get("ajustes_por_nota", {}).get(nota_normalizada, 0))
        ajuste_descricao_nota = int(
            contexto_recalibracao.get("ajustes_por_descricao_nota", {}).get(
                (descricao_normalizada, nota_normalizada),
                0,
            )
        )

        ajuste_total = ajuste_nota + ajuste_descricao_nota
        ajuste_total = max(AJUSTE_MAXIMO_NEGATIVO, min(AJUSTE_MAXIMO_POSITIVO, ajuste_total))

        detalhes: list[str] = []
        if ajuste_nota > 0:
            detalhes.append("nota reforcada")
        elif ajuste_nota < 0:
            detalhes.append("nota penalizada")

        if ajuste_descricao_nota > 0:
            detalhes.append("descricao reforcada")
        elif ajuste_descricao_nota < 0:
            detalhes.append("descricao penalizada")

        return ajuste_total, ", ".join(detalhes) or "sem ajuste"

    def _estimar_metricas_recalibradas(
        self,
        estados_feedback: list[tuple[FeedbackSugestaoNota, str]],
        contexto_recalibracao: dict[str, Any],
    ) -> dict[str, Any]:
        """Estima o efeito da recalibracao sobre o feedback ja recolhido."""
        total_analisadas = len(estados_feedback)
        total_com_sugestao_recalibrada = 0
        total_aceites_recalibrado = 0
        total_rejeitadas_recalibrado = 0

        for feedback, estado in estados_feedback:
            if not feedback.sugestao_1.strip():
                continue

            ajuste, _ = self._calcular_ajuste_recalibracao(
                descricao=feedback.descricao,
                nota=feedback.sugestao_1,
                contexto_recalibracao=contexto_recalibracao,
            )
            score_estimado = max(0, min(100, int(feedback.score_1 or 0) + ajuste))
            if score_estimado < LIMIAR_SUGESTAO_RECALIBRADA:
                continue

            total_com_sugestao_recalibrada += 1
            if estado == "ACEITE":
                total_aceites_recalibrado += 1
            elif estado == "REJEITADA":
                total_rejeitadas_recalibrado += 1

        rotulos_notas = contexto_recalibracao.get("rotulos_notas", {})
        ajustes_por_nota = contexto_recalibracao.get("ajustes_por_nota", {})

        top_penalizadas = sorted(
            (
                (rotulos_notas.get(chave, chave), ajuste)
                for chave, ajuste in ajustes_por_nota.items()
                if ajuste < 0
            ),
            key=lambda item: item[1],
        )[:5]
        top_reforcadas = sorted(
            (
                (rotulos_notas.get(chave, chave), ajuste)
                for chave, ajuste in ajustes_por_nota.items()
                if ajuste > 0
            ),
            key=lambda item: item[1],
            reverse=True,
        )[:5]

        return {
            "cobertura_recalibrada_estimada": self._calcular_percentagem(
                total_com_sugestao_recalibrada,
                total_analisadas,
            ),
            "aceitacao_recalibrada_estimada": self._calcular_percentagem(
                total_aceites_recalibrado,
                total_com_sugestao_recalibrada,
            ),
            "rejeicao_recalibrada_estimada": self._calcular_percentagem(
                total_rejeitadas_recalibrado,
                total_com_sugestao_recalibrada,
            ),
            "total_padroes_recalibrados": len(contexto_recalibracao.get("ajustes_por_nota", {}))
            + len(contexto_recalibracao.get("ajustes_por_descricao_nota", {})),
            "top_sugestoes_penalizadas": top_penalizadas,
            "top_sugestoes_reforcadas": top_reforcadas,
        }

    @staticmethod
    def _criar_estatisticas_feedback() -> dict[str, int]:
        """Cria um bloco simples de contadores para um padrao."""
        return {
            "aceites": 0,
            "rejeitadas": 0,
            "editadas": 0,
            "respostas": 0,
        }

    @staticmethod
    def _acumular_estado_estatistico(estatisticas: dict[str, int], estado: str) -> None:
        """Acumula o estado normalizado num bloco de estatisticas."""
        if estado == "ACEITE":
            estatisticas["aceites"] += 1
        elif estado == "REJEITADA":
            estatisticas["rejeitadas"] += 1
        elif estado == "EDITADA":
            estatisticas["editadas"] += 1

        if estado in {"ACEITE", "REJEITADA", "EDITADA"}:
            estatisticas["respostas"] += 1

    @staticmethod
    def _calcular_ajuste_estatistico(
        estatisticas: dict[str, int],
        minimo_respostas: int,
        peso_aceite: int,
        peso_rejeitada: int,
        peso_editada: int,
        bonus_aceite: int,
        bonus_falha: int,
        limite_inferior: int,
        limite_superior: int,
    ) -> int:
        """Converte contagens de feedback num ajuste pequeno e explicavel."""
        respostas = int(estatisticas.get("respostas", 0))
        if respostas < minimo_respostas:
            return 0

        aceites = int(estatisticas.get("aceites", 0))
        rejeitadas = int(estatisticas.get("rejeitadas", 0))
        editadas = int(estatisticas.get("editadas", 0))

        ajuste = (aceites * peso_aceite) - (rejeitadas * peso_rejeitada) - (editadas * peso_editada)
        taxa_aceite = aceites / respostas if respostas else 0
        taxa_falha = (rejeitadas + editadas) / respostas if respostas else 0

        if taxa_aceite >= 0.70:
            ajuste += bonus_aceite
        if taxa_falha >= 0.60:
            ajuste -= bonus_falha

        return max(limite_inferior, min(limite_superior, ajuste))

    @staticmethod
    def _calcular_percentagem(parte: int, total: int) -> float:
        """Calcula percentagem simples com duas casas."""
        if total <= 0:
            return 0.0

        return round((parte / total) * 100, 2)
